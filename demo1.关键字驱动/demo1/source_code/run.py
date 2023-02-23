import openpyxl
from demo1.source_code.action_page import KeyWords
from datetime import datetime
import traceback
import time
from multiprocessing.dummy import Pool
import os


class RunTest:
    def runTest(self,filename):
        """
        :param filename:
        :return:
        """
        #读取自动化测试脚本文件
        fp = openpyxl.load_workbook(filename)
        # 获取sheet_names
        for sheetname in fp.sheetnames:
            # sheet_name 是用例sheet的标签名，用于后面的错误截图
            save_name = sheetname
            ws = fp[sheetname]      #读取每个sheet的内容
            # 获取最大行数
            ws_rows_max = ws.max_row
            idx = 3
            data = {}       #定义一个字典，获取读取到的关键字（keyword），定位方式（element），定位元素（name），参数（value）
            demo = KeyWords()       #实例化关键字的对象
            while idx <= ws_rows_max:
                keyword = ws.cell(row=idx,column=3).value       #获取关键字
                element = ws.cell(row=idx,column=4).value       #获取元素定位类型
                name = ws.cell(row=idx,column=5).value          #获取定位信息
                value = ws.cell(row=idx,column=6).value         #获取输入参数
                image_name = ws.cell(row=idx,column=2).value    #获取操作步骤
                data['element'] = element
                data['name'] = name
                data['value'] = value
                # data['save_name'] = save_name+image_name
                #删除字典value值为None的键值对data
                for key in list(data.keys()):
                    if data.get(key) == None:
                        del data[key]
                try:
                    status = getattr(demo,keyword)(**data)
                    #在excel的测试执行时间列写入执行时间
                    ws.cell(row=idx,column=7).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    if str(keyword).startswith('assert'):
                        if status == 1:      # 如果断言无误则将执行结果写入Passed
                            ws.cell(row=idx, column=8).value = 'Passed'
                        elif status == 0:       # 如果断言错误则将执行结果写入Failed
                            ws.cell(row=idx, column=8).value = 'Failed'
                            demo.save_image(save_name+image_name)       #并保存截图用sheet页名称+执行步骤名称命名
                        fp.save(filename)
                    elif keyword == 'repeat':
                        i = 0
                        while i < value:
                            self.runTest(filename)
                            i += 1
                    else:
                        # 如果执行无误则将执行结果写入Pass
                        ws.cell(row=idx, column=8).value = 'pass'
                        fp.save(filename)
                except Exception:
                        # 如果执行出现异常，测试结果写入 异常，同时将异常信息写入到表格中
                        ws.cell(row=idx, column=8).value = '异常'
                        ws.cell(row=idx, column=9).value = traceback.format_exc()
                        demo.save_image(save_name+image_name)
                idx += 1

    def getExcel(self,path_name='../test_case/'):
        """
        :param path_name: path_name:自动化excel文件的目录
        :return:
        """
        # 获取 test_case 文件夹下的所有 excel 文件
        file_list = []
        for el in os.listdir(path_name):
            excel_path = path_name + el
            file_list.append(excel_path)
        return file_list

    def runPool(self,func,iterable,threads_num=1):
        """
        :param func: 执行名称
        :param iterable: 迭代对象
        :param pool_num: 开启的进程组数，默认开启一个线程组
        :return:
        """
        start = time.time()
        pool = Pool(threads_num)
        pool.map(func=func,iterable=iterable)
        pool.close()
        pool.join()
        print(time.time()-start)


if __name__ == '__main__':
    runner = RunTest()
    filenames = runner.getExcel()
    # for filename in filenames:
    #     runner.runTest(filename)
    runner.runPool(runner.runTest,filenames)
